"""
NOVATECH.MA — ETL: Synchronisation Stock OneDrive → Supabase
Fichier: etl/sync_stock.py

Usage:
  python sync_stock.py                    # sync depuis OneDrive
  python sync_stock.py --local fichier.xlsx  # sync depuis fichier local (dev)

Cron (quotidien 7h00):
  0 7 * * 1-6 /home/ubuntu/novatech/etl/cron.sh
"""

import os
import sys
import io
import re
import time
import logging
import argparse
from datetime import datetime
from typing import Optional

import openpyxl
import requests
from supabase import create_client, Client
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("/var/log/novatech_sync.log", mode="a"),
    ]
)
log = logging.getLogger(__name__)

# ─── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL  = os.environ["SUPABASE_URL"]
SUPABASE_KEY  = os.environ["SUPABASE_SERVICE_KEY"]
ONEDRIVE_URL  = os.environ.get("ONEDRIVE_STOCK_URL", "")
STOCK_SHEET   = "             Stock             "

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ─── Mapping familles → catégories ─────────────────────────────────────────────
FAMILY_MAP = {
    "DELL PRO MICRO":     {"cat": "ordinateurs", "sub": "mini-pc",         "brand": "Dell"},
    "DELL PC 7020SFF":    {"cat": "ordinateurs", "sub": "pc-bureau",       "brand": "Dell"},
    "DELL PC 7020MFF":    {"cat": "ordinateurs", "sub": "mini-pc",         "brand": "Dell"},
    "DELL PC 7420":       {"cat": "ordinateurs", "sub": "tout-en-un",      "brand": "Dell"},
    "DELL LAPTOP 3450":   {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Dell"},
    "DELL LAPTOP 7350":   {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Dell"},
    "DELL PRO 13 PREM":   {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Dell"},
    "DELL PRO 13 PLUS":   {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Dell"},
    "DELL PRO 14 PLUS":   {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Dell"},
    "DELL PRO 14 PREM":   {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Dell"},
    "DELL LAPTOP XPS":    {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Dell"},
    "LENOVO LAPTOP":      {"cat": "ordinateurs", "sub": "pc-portable",     "brand": "Lenovo"},
    "DELL ECRAN":         {"cat": "ecrans",       "sub": "moniteurs",      "brand": "Dell"},
    "DELL ECRAN ":        {"cat": "ecrans",       "sub": "moniteurs",      "brand": "Dell"},
    "TV Philips":         {"cat": "ecrans",       "sub": "signage",        "brand": "Philips"},
    "VIDEOPROJECTEUR":    {"cat": "ecrans",       "sub": "videoprojecteurs","brand": "Multi"},
    "DELL DOCK":          {"cat": "accessoires",  "sub": "docking",        "brand": "Dell"},
    "DELL KYB & MOUSE":   {"cat": "accessoires",  "sub": "claviers-souris","brand": "Dell"},
    "DELL LAPTOP ACC":    {"cat": "accessoires",  "sub": "adaptateurs",    "brand": "Dell"},
    "DELL ECRAN ACC":     {"cat": "accessoires",  "sub": "adaptateurs",    "brand": "Dell"},
    "ACCESSOIRE":         {"cat": "accessoires",  "sub": "divers",         "brand": "Multi"},
    "DOUCHETTE":          {"cat": "accessoires",  "sub": "scanners",       "brand": "Multi"},
    "DELL BP & CASE":     {"cat": "bagagerie",    "sub": "sacoches",       "brand": "Dell"},
    "BP & CASE":          {"cat": "bagagerie",    "sub": "sacoches",       "brand": "Multi"},
    "JABRA CASQUE":       {"cat": "audio",        "sub": "casques-pro",    "brand": "Jabra"},
    "JABRA ACC":          {"cat": "audio",        "sub": "accessoires-audio","brand": "Jabra"},
    "JABRA ACC+CASQUE":   {"cat": "audio",        "sub": "casques-pro",    "brand": "Jabra"},
    "POLY CASQUE":        {"cat": "audio",        "sub": "casques-pro",    "brand": "Poly"},
    "POLY ACC":           {"cat": "audio",        "sub": "accessoires-audio","brand": "Poly"},
    "DELL CASQUE":        {"cat": "audio",        "sub": "casques-pro",    "brand": "Dell"},
    "IMPRIMANTE & CONSO": {"cat": "impression",   "sub": "imprimantes",    "brand": "Multi"},
    "CISCO":              {"cat": "reseau",        "sub": "switches",       "brand": "Cisco"},
    "CISCO ACC":          {"cat": "reseau",        "sub": "accessoires-reseau","brand": "Cisco"},
    "Ubiquiti":           {"cat": "reseau",        "sub": "wifi",           "brand": "Ubiquiti"},
    "ROUTER":             {"cat": "reseau",        "sub": "routeurs",       "brand": "Multi"},
}
DEFAULT_FAMILY = {"cat": "accessoires", "sub": "divers", "brand": "Multi"}


def slugify(text: str) -> str:
    t = text.lower()
    for src, dst in [("à","a"),("â","a"),("é","e"),("è","e"),("ê","e"),
                     ("î","i"),("ô","o"),("û","u"),("ù","u"),("ç","c")]:
        t = t.replace(src, dst)
    t = re.sub(r"[^a-z0-9\s\-]", "", t)
    t = re.sub(r"\s+", "-", t.strip())
    t = re.sub(r"-{2,}", "-", t)
    return t[:200]


def make_commercial_title(raw: str, brand: str) -> str:
    if not raw:
        return brand
    clean = raw.strip()
    # supprimer préfixe code interne type "AB1234 "
    clean = re.sub(r"^[A-Z]{1,3}[0-9]{3,8}\s+", "", clean)
    return clean[:200]


def make_slug_unique(base: str, existing: set) -> str:
    slug = base
    counter = 2
    while slug in existing:
        slug = f"{base}-{counter}"
        counter += 1
    return slug


def get_dell_image_url(sku: str) -> Optional[str]:
    """URL image officielle Dell (CDN public)"""
    return f"https://i.dell.com/is/image/DellContent/{sku}?fmt=jpg&wid=500&hei=500"


def get_image_url(sku: str, brand: str) -> Optional[str]:
    if brand.lower() == "dell":
        return get_dell_image_url(sku)
    # Pour les autres marques, sera enrichi manuellement ou via admin
    return None


def to_int(val) -> int:
    try:
        return int(float(str(val))) if val is not None else 0
    except:
        return 0


def download_from_onedrive(url: str) -> bytes:
    # Convertir URL OneDrive partagée en URL de download direct
    dl_url = url
    if "1drv.ms" in url or "sharepoint.com" in url:
        if "?" in url:
            dl_url = url + "&download=1"
        else:
            dl_url = url + "?download=1"
    r = requests.get(dl_url, allow_redirects=True, timeout=60)
    r.raise_for_status()
    return r.content


def load_workbook_bytes(content: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)


def sync(local_path: str = None):
    started = time.time()
    stats = {"processed": 0, "imported": 0, "updated": 0, "skipped": 0, "errors": 0, "error_details": []}

    try:
        # ── 1. Charger le fichier ──────────────────────────────────────────
        if local_path:
            log.info(f"Mode local: {local_path}")
            with open(local_path, "rb") as f:
                content = f.read()
        else:
            log.info("Téléchargement depuis OneDrive...")
            content = download_from_onedrive(ONEDRIVE_URL)
        log.info(f"Fichier chargé ({len(content)/1024:.1f} KB)")

        wb = load_workbook_bytes(content)

        # Trouver l'onglet Stock
        sheet_name = None
        for name in wb.sheetnames:
            if "Stock" in name:
                sheet_name = name
                break
        if not sheet_name:
            raise ValueError(f"Onglet Stock non trouvé. Onglets: {wb.sheetnames}")

        ws = wb[sheet_name]
        log.info(f"Onglet trouvé: '{sheet_name}'")

        # ── 2. Récupérer slugs existants ───────────────────────────────────
        existing_slugs_res = supabase.table("products").select("sku,slug,id").execute()
        existing_map = {r["sku"]: r for r in (existing_slugs_res.data or [])}
        used_slugs   = {r["slug"] for r in (existing_slugs_res.data or [])}

        # ── 3. Parser les lignes ───────────────────────────────────────────
        for row in ws.iter_rows(min_row=5, max_row=500, values_only=True):
            if not row or len(row) < 5:
                continue
            zone, famille, sku, designation, stock_qty, stock_incoming = (
                row[0], row[1], row[2], row[3],
                row[4] if len(row) > 4 else 0,
                row[5] if len(row) > 5 else 0,
            )
            date_ctrl = row[6] if len(row) > 6 else None
            obs        = row[7] if len(row) > 7 else None

            if not sku:
                continue
            sku = str(sku).strip()
            if not sku or sku.lower() in ("sku","ref","référence",""):
                continue

            stats["processed"] += 1
            famille_clean = str(famille).strip() if famille else ""
            meta          = FAMILY_MAP.get(famille_clean, DEFAULT_FAMILY)
            sq            = to_int(stock_qty)
            si            = to_int(stock_incoming)

            try:
                if sku in existing_map:
                    # ── UPDATE ────────────────────────────────────────────
                    supabase.table("products").update({
                        "stock_qty":       sq,
                        "stock_incoming":  si,
                        "last_checked_at": str(date_ctrl) if date_ctrl else None,
                        "internal_note":   str(obs) if obs else None,
                        "last_import_at":  datetime.now().isoformat(),
                    }).eq("sku", sku).execute()
                    stats["updated"] += 1

                else:
                    # ── INSERT ────────────────────────────────────────────
                    title = make_commercial_title(str(designation) if designation else "", meta["brand"])
                    slug  = make_slug_unique(slugify(f"{meta['brand']} {title}"), used_slugs)
                    used_slugs.add(slug)
                    image = get_image_url(sku, meta["brand"])

                    supabase.table("products").insert({
                        "sku":             sku,
                        "brand":           meta["brand"],
                        "category":        meta["cat"],
                        "subcategory":     meta["sub"],
                        "family_raw":      famille_clean,
                        "warehouse_zone":  str(zone) if zone else None,
                        "title_raw":       str(designation) if designation else None,
                        "title_commercial":title,
                        "slug":            slug,
                        "stock_qty":       sq,
                        "stock_incoming":  si,
                        "last_checked_at": str(date_ctrl) if date_ctrl else None,
                        "internal_note":   str(obs) if obs else None,
                        "image_main":      image,
                        "image_source":    "dell_api" if meta["brand"] == "Dell" else None,
                        "status":          "draft",
                        "is_published":    False,
                        "last_import_at":  datetime.now().isoformat(),
                    }).execute()
                    stats["imported"] += 1

            except Exception as e:
                stats["errors"] += 1
                stats["error_details"].append({"sku": sku, "error": str(e)})
                log.error(f"  ✗ SKU {sku}: {e}")

        # ── 4. Logger ──────────────────────────────────────────────────────
        duration_ms = int((time.time() - started) * 1000)
        supabase.table("import_logs").insert({
            "filename":       local_path or "STOCK_PHYSIQUE.xlsx (OneDrive)",
            "source":         "local" if local_path else "onedrive",
            "rows_processed": stats["processed"],
            "imported":       stats["imported"],
            "updated":        stats["updated"],
            "errors":         stats["errors"],
            "error_details":  stats["error_details"][:50],
            "status":         "success" if stats["errors"] == 0 else "partial",
            "duration_ms":    duration_ms,
        }).execute()

        log.info(
            f"✅ Sync terminée en {duration_ms}ms — "
            f"{stats['processed']} traités / "
            f"{stats['imported']} nouveaux / "
            f"{stats['updated']} mis à jour / "
            f"{stats['errors']} erreurs"
        )

    except Exception as fatal:
        log.error(f"❌ Erreur fatale: {fatal}")
        try:
            supabase.table("import_logs").insert({
                "status":        "error",
                "error_details": [{"error": str(fatal)}],
            }).execute()
        except:
            pass
        sys.exit(1)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--local", help="Chemin vers fichier Excel local (optionnel)")
    args = parser.parse_args()
    sync(local_path=args.local)
